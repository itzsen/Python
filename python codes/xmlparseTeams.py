import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET

filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

username='ucadmin'
password='Syr!nx2112'
ip='10.2.2.35'


def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()



teamlayout='http://'+ip+'/finesse/api/Teams'


# GET teamslayout
response=GET(teamlayout,username,password)
xmlcontent='teamlayout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('teamlayout.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Teams')
sheetmaxrow=sheet.max_row+1
sheetmaxrow1=sheetmaxrow
sheetmaxrow2=sheetmaxrow
sheetmaxrow3=sheetmaxrow
sheetmaxrow4=sheetmaxrow
sheetmaxrow5=sheetmaxrow
sheetmaxrow6=sheetmaxrow

for child in root.findall('./Team/uri'):
    teamuri=child.text
    teamurivalue=teamuri.split('/')
    teamurl=('http://'+ip+'/finesse/api/Team/'+teamurivalue[4])
    response=GET(teamurl,username,password)
    xmlteamfile=('team'+teamurivalue[4])+'.xml'
    fileopen=open(xmlteamfile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlteamfile)
    subroot = subtree.getroot()

    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('id'):
        sheet.cell(sheetmaxrow,2,subchild.text)
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,3,subchild.text)

    #team URLs
    tnotreadyurl=('http://'+ip+'/finesse/api/Team/'+teamurivalue[4]+'/ReasonCodes?category=NOT_READY')
    tlogouturl=('http://'+ip+'/finesse/api/Team/'+teamurivalue[4]+'/ReasonCodes?category=LOGOUT')
    twrapurl=('http://'+ip+'/finesse/api/Team/'+teamurivalue[4]+'/WrapUpReasons')
    tphoneurl=('http://'+ip+'/finesse/api/Team/'+teamurivalue[4]+'/PhoneBooks')
    tlayouturl=('http://'+ip+'/finesse/api/Team/'+teamurivalue[4]+'/LayoutConfig')
    tworkflowurl=('http://'+ip+'/finesse/api/Team/'+teamurivalue[4]+'/Workflows')

    #GET team notready              
    response=GET(tnotreadyurl,username,password)
    tnotreadyxmlcontent='tnotready'+teamurivalue[4]+'.xml'
    fileopen=open(tnotreadyxmlcontent,'wb')
    fileopen.write(response.content)
    fileopen.close()

    sub1tree= ET.parse(tnotreadyxmlcontent)
    sub1root = sub1tree.getroot()

    for RC in sub1root.iter('ReasonCode'):
        for label in RC.findall('label'):
            sheet.cell(sheetmaxrow1,4,label.text)
            sheetmaxrow1=sheetmaxrow1+1

    #GET team logout              
    response=GET(tlogouturl,username,password)
    tlogoutxmlcontent='tlogout'+teamurivalue[4]+'.xml'
    fileopen=open(tlogoutxmlcontent,'wb')
    fileopen.write(response.content)
    fileopen.close()
    
    sub2tree= ET.parse(tlogoutxmlcontent)
    sub2root = sub2tree.getroot()

    for RC in sub2root.iter('ReasonCode'):
        for label in RC.findall('label'):
            sheet.cell(sheetmaxrow2,5,label.text)
            sheetmaxrow2=sheetmaxrow2+1

    #GET team wrapup              
    response=GET(twrapurl,username,password)
    twrapxmlcontent='twrap'+teamurivalue[4]+'.xml'
    fileopen=open(twrapxmlcontent,'wb')
    fileopen.write(response.content)
    fileopen.close()
    
    sub3tree= ET.parse(twrapxmlcontent)
    sub3root = sub3tree.getroot()

    for RC in sub3root.iter('WrapUpReason'):
        for label in RC.findall('label'):
            sheet.cell(sheetmaxrow3,6,label.text)
            sheetmaxrow3=sheetmaxrow3+1    
        
    #GET team phonebook              
    response=GET(tphoneurl,username,password)
    tphonexmlcontent='tphone'+teamurivalue[4]+'.xml'
    fileopen=open(tphonexmlcontent,'wb')
    fileopen.write(response.content)
    fileopen.close()

    sub4tree= ET.parse(tphonexmlcontent)
    sub4root = sub4tree.getroot()

    for RC in sub4root.iter('PhoneBook'):
        for label in RC.findall('name'):
            sheet.cell(sheetmaxrow4,7,label.text)
            sheetmaxrow4=sheetmaxrow4+1    

    #GET team workflow              
    response=GET(tworkflowurl,username,password)
    twfxmlcontent='twf'+teamurivalue[4]+'.xml'
    fileopen=open(twfxmlcontent,'wb')
    fileopen.write(response.content)
    fileopen.close()

    sub5tree= ET.parse(twfxmlcontent)
    sub5root = sub5tree.getroot()

    for RC in sub5root.iter('Workflow'):
        for label in RC.findall('name'):
            sheet.cell(sheetmaxrow5,8,label.text)
            sheetmaxrow5=sheetmaxrow5+1    

    #GET team layout              
    response=GET(tlayouturl,username,password)
    tlayoutxmlcontent='tlayout'+teamurivalue[4]+'.xml'
    fileopen=open(tlayoutxmlcontent,'wb')
    fileopen.write(response.content)
    fileopen.close()

    sub6tree= ET.parse(tlayoutxmlcontent)
    sub6root = sub6tree.getroot()

    for RC in sub6root.findall('useDefault'):
        if RC.text=='true':
            sheet.cell(sheetmaxrow6,9,'Default')
        else:
            for RC in sub6root.findall('layoutxml'):
                sheet.cell(sheetmaxrow6,9,RC.text)  

    sheetmaxrow=sheet.max_row+1
    sheetmaxrow1=sheetmaxrow
    sheetmaxrow2=sheetmaxrow
    sheetmaxrow3=sheetmaxrow
    sheetmaxrow4=sheetmaxrow
    sheetmaxrow5=sheetmaxrow
    sheetmaxrow6=sheetmaxrow 

wb.save('Finesse.xlsx')

'''

# Clear all temporary files

files=os.listdir(os.getcwd())

for file in files:
    if file.endswith('.xml'):
        os.remove(file)


'''
