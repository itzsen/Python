import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill,Border,Fill
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Workflow')




blackfill= PatternFill(start_color='FF000000',
                       end_color='FF000000', fill_type='solid')



lastrow=sheet.max_row

for cell in range(6,lastrow+1):
    sheet['N'+str(cell)].fill=blackfill

wb.save('Finesse.xlsx')


'''
os.chdir('C:\\Users\\senthil.natarajan\\Dropbox\\Programming\\Python')

for filenum in range (1,100):
    if os.path.isfile('workflowaction'+str(filenum)+'.xml'):                 
        wfactionfile='workflowaction'+str(filenum)+'.xml'
        print (wfactionfile +' exist')
    else:
        continue

'''



