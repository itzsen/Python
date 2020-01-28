import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET


filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

username='ucadmin'
password='Syr!nx2112'
ip='10.2.2.35'


def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()



wflayout='http://'+ip+'/finesse/api/Workflows'
wfAlayout='http://'+ip+'/finesse/api/WorkflowActions'

# GET workflowactionlayout
response=GET(wfAlayout,username,password)
xmlcontent='wfAlayout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('wfAlayout.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Workflow')
sheetmaxrow=sheet.max_row+1
sheetmaxrow1=sheetmaxrow

for child in root.findall('./WorkflowAction/uri'):
    wfuri=child.text
    wfurivalue=wfuri.split('/')
    wfurl=('http://'+ip+'/finesse/api/WorkflowAction/'+wfurivalue[4])
    response=GET(wfurl,username,password)
    xmlwffile=('workflowAction'+wfurivalue[4])+'.xml'
    fileopen=open(xmlwffile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlwffile)
    subroot = subtree.getroot()

    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow1,12,subchild.text)
    for subchild in subroot.findall('type'):
        sheet.cell(sheetmaxrow1,13,subchild.text)
    for subchild in subroot.findall('handledBy'):
        sheet.cell(sheetmaxrow1,14,subchild.text)        
    sheetmaxrow1=sheetmaxrow1+1


sheetmaxrow=sheet.max_row+1
wb.save('Finesse.xlsx')


'''
    subtree= ET.parse(xmlwffile)
    subroot = subtree.getroot()

    for subchild in subroot.findall('.//workflowActions//uri'):
        sheet.cell(sheetmaxrow7,12,subchild.text)
        sheetmaxrow7=sheetmaxrow7+1
        #GET workflow Action Parameters
        wfAuri='http://'+ip+subchild.text
        wfAurivalue=wfAuri.split('/')
        wfAurl=('http://'+ip+'/finesse/api/WorkflowAction/'+wfurivalue[4])
        response=GET(wfAurl,username,password)
        wfAfile=('workflowaction'+wfAurivalue[6])+'.xml'
        fileopen=open(wfAfile,'wb')
        fileopen.write(response.content)
        fileopen.close()
        sub1tree= ET.parse(wfAfile)
        sub1root = sub1tree.getroot()      
        for sub1child in sub1root.findall('handledBy'):
            sheet.cell(sheetmaxrow8,14,sub1child.text)
            sheetmaxrow8=sheetmaxrow8+1

sheetmaxrow1=sheetmaxrow
sheetmaxrow2=sheetmaxrow
sheetmaxrow3=sheetmaxrow
sheetmaxrow4=sheetmaxrow
sheetmaxrow5=sheetmaxrow
sheetmaxrow6=sheetmaxrow
sheetmaxrow7=sheetmaxrow
sheetmaxrow8=sheetmaxrow
sheetmaxrow9=sheetmaxrow
sheetmaxrow10=sheetmaxrow



    sheetmaxrow=sheet.max_row+1
    
wb.save('Finesse.xlsx')


=====





    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('description'):
        sheet.cell(sheetmaxrow,2,subchild.text)
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,3,subchild.text)
    for subchild in subroot.findall('./TriggerSet/type'):
        sheet.cell(sheetmaxrow,4,subchild.text)
    for subchild in subroot.findall('./TriggerSet/name'):
        sheet.cell(sheetmaxrow,5,subchild.text)
    for subchild in subroot.findall('./ConditionSet/applyMethod'):
        sheet.cell(sheetmaxrow,6,subchild.text)
    for subchild in subroot.findall('.//ConditionSet//type'):
        sheet.cell(sheetmaxrow1,7,subchild.text)
        sheetmaxrow1=sheetmaxrow1+1
    for subchild in subroot.findall('.//ConditionSet//comparator'):
        sheet.cell(sheetmaxrow2,9,subchild.text)
        sheetmaxrow2=sheetmaxrow2+1   
    for subchild in subroot.findall('.//ConditionSet//value'):
        sheet.cell(sheetmaxrow3,10,subchild.text)
        sheetmaxrow3=sheetmaxrow3+1
    for subchild in subroot.findall('.//ConditionSet//name'):
        sheet.cell(sheetmaxrow4,8,subchild.text)
        sheetmaxrow4=sheetmaxrow4+1
    for subchild in subroot.findall('.//workflowActions//name'):
        sheet.cell(sheetmaxrow5,11,subchild.text)
        sheetmaxrow5=sheetmaxrow5+1
    for subchild in subroot.findall('.//workflowActions//type'):
        sheet.cell(sheetmaxrow6,13,subchild.text)
        sheetmaxrow6=sheetmaxrow6+1







======


        for sub1child in sub1root.findall('.//params//value'):
            sheet.cell(sheetmaxrow9,15,sub1child.text)
            sheetmaxrow9=sheetmaxrow9+1




        for sub1child in sub1root.findall('.//value'):
            sheet.cell(sheetmaxrow5,16,sub1child.text)
            sheetmaxrow5=sheetmaxrow5+1


        for sub1child in sub1root.findall('.//params//Param//value'):
            sheet.cell(sheetmaxrow5,15,subchild.text)
            sheetmaxrow5=sheetmaxrow5+1
'''

