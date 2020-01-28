import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET


filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

username='ucadmin'
password='Syr!nx2112'

def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()



wrapup='http://10.2.2.35/finesse/api/WrapUpReasons'

# GET not-ready

response=GET(wrapup,username,password)
xmlcontent='wrapup.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()
i=7
j=7
k=7
tree= ET.parse('wrapup.xml')
root = tree.getroot()
for child in root:
    for element in child.findall('uri'):
        wb=openpyxl.load_workbook('Finesse.xlsx')
        sheet=wb.get_sheet_by_name('Wrapup')
        sheet.cell(i,3,element.text)
        i=i+1
        wb.save('Finesse.xlsx')
    for element in child.findall('label'):
        wb=openpyxl.load_workbook('Finesse.xlsx')
        sheet=wb.get_sheet_by_name('Wrapup')
        sheet.cell(j,1,element.text)
        j=j+1
        wb.save('Finesse.xlsx')
    for element in child.findall('forAll'):
        wb=openpyxl.load_workbook('Finesse.xlsx')
        sheet=wb.get_sheet_by_name('Wrapup')
        sheet.cell(k,2,element.text)
        k=k+1
        wb.save('Finesse.xlsx')
