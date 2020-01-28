import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET


filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

username='ucadmin'
password='Syr!nx2112'
ip='10.2.2.35'


def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()



phonelayout='http://'+ip+'/finesse/api/PhoneBooks'
contactlayout=''

# GET phonelayout
response=GET(phonelayout,username,password)
xmlcontent='phonelayout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('phonelayout.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Phonebook')
sheetmaxrow=sheet.max_row+1

for child in root.findall('.//PhoneBook//uri'):
    phonebookuri=child.text
    phonebookurivalue=phonebookuri.split('/')
    phonebookurl=('http://'+ip+'/finesse/api/PhoneBook/'+phonebookurivalue[4])
    response=GET(phonebookurl,username,password)
    xmlcontentfile=('phonebook'+phonebookurivalue[4])+'.xml'
    fileopen=open(xmlcontentfile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlcontentfile)
    subroot = subtree.getroot()
    for subchild in subroot.findall('type'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,3,subchild.text)
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,4,subchild.text)
    for subchild in subroot.findall('contacts'):
        sheet.cell(sheetmaxrow,2,subchild.text)
        contactlayout='http://'+ip+subchild.text
        # GET contactlayout
        response=GET(contactlayout,username,password)
        cxmlcontent='contactlayout.xml'
        fileopen=open(cxmlcontent,'wb')
        fileopen.write(response.content)
        fileopen.close()

        ctree= ET.parse('contactlayout.xml')
        croot = ctree.getroot()
        sheetmaxrow1=sheetmaxrow
        sheetmaxrow2=sheetmaxrow
        sheetmaxrow3=sheetmaxrow
        sheetmaxrow4=sheetmaxrow
        sheetmaxrow5=sheetmaxrow
        for cchild in croot.findall('.//Contact//description'):
            sheet.cell(sheetmaxrow1,8,cchild.text)
            sheetmaxrow1=sheetmaxrow1+1
        for cchild in croot.findall('.//Contact//firstName'):
            sheet.cell(sheetmaxrow2,5,cchild.text)
            sheetmaxrow2=sheetmaxrow2+1
        for cchild in croot.findall('.//Contact//lastName'):
            sheet.cell(sheetmaxrow3,6,cchild.text)
            sheetmaxrow3=sheetmaxrow3+1
        for cchild in croot.findall('.//Contact//phoneNumber'):
            sheet.cell(sheetmaxrow4,7,cchild.text)
            sheetmaxrow4=sheetmaxrow4+1
        for cchild in croot.findall('.//Contact//uri'):
            sheet.cell(sheetmaxrow5,9,cchild.text)
            sheetmaxrow5=sheetmaxrow5+1

    sheetmaxrow=sheet.max_row+1
    
wb.save('Finesse.xlsx')


