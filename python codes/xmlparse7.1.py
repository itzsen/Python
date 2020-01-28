import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET

filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

username='ucadmin'
password='Syr!nx2112'
ip='10.2.2.35'


def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()



wflayout='http://'+ip+'/finesse/api/Workflows'
wfAlayout=''

# GET workflowlayout
response=GET(wflayout,username,password)
xmlcontent='wflayout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('wflayout.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Workflow')
sheetmaxrow=sheet.max_row+1
sheetmaxrow1=sheetmaxrow
sheetmaxrow2=sheetmaxrow
sheetmaxrow3=sheetmaxrow


for child in root.findall('./Workflow/uri'):
    wfuri=child.text
    wfurivalue=wfuri.split('/')
    wfurl=('http://'+ip+'/finesse/api/Workflow/'+wfurivalue[4])
    response=GET(wfurl,username,password)
    xmlwffile=('workflow'+wfurivalue[4])+'.xml'
    fileopen=open(xmlwffile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlwffile)
    subroot = subtree.getroot()

    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('description'):
        sheet.cell(sheetmaxrow,2,subchild.text)
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,3,subchild.text)
    for subchild in subroot.findall('./TriggerSet/type'):
        sheet.cell(sheetmaxrow,4,subchild.text)
    for subchild in subroot.findall('./TriggerSet/name'):
        sheet.cell(sheetmaxrow,5,subchild.text)
    for subchild in subroot.findall('./ConditionSet/applyMethod'):
        sheet.cell(sheetmaxrow,6,subchild.text)
    for Condition in subroot.iter('Condition'):             
        for typec in Condition.iter('type'):
            sheet.cell(sheetmaxrow1,7,typec.text)
        for comparec in Condition.iter('comparator'):
            sheet.cell(sheetmaxrow1,9,comparec.text)
        for namec in Condition.iter('name'):
            sheet.cell(sheetmaxrow1,8,namec.text)
        for valuec in Condition.iter('value'):
            sheet.cell(sheetmaxrow1,10,valuec.text)      
        sheetmaxrow1=sheet.max_row+1

    for wfa in subroot.iter('WorkflowAction'):
        for wname in wfa.iter('name'):
            sheet.cell(sheetmaxrow2,11,wname.text)
        for wuri in wfa.iter('uri'):
            sheet.cell(sheetmaxrow2,12,wuri.text)         
        for wtype in wfa.iter('type'):
            sheet.cell(sheetmaxrow2,13,wtype.text) 
        sheetmaxrow2=sheetmaxrow2+1

    for subchild in subroot.findall('.//workflowActions//uri'):
        #GET workflow Action Parameters
        wfAuri='http://'+ip+subchild.text
        wfAurivalue=wfAuri.split('/')
        wfAurl=('http://'+ip+'/finesse/api/WorkflowAction/'+wfAurivalue[6])
        response=GET(wfAurl,username,password)
        wfAfile=('workflowaction'+wfAurivalue[6])+'.xml'
        fileopen=open(wfAfile,'wb')
        fileopen.write(response.content)
        fileopen.close()
   
    sheetmaxrow=sheet.max_row+2
    sheetmaxrow1=sheetmaxrow
    sheetmaxrow2=sheetmaxrow

for filenum in range (1,100):
    if os.path.isfile('workflowaction'+str(filenum)+'.xml'):
        wfactionfile='workflowaction'+str(filenum)+'.xml'
        sub1tree= ET.parse(wfactionfile)
        sub1root = sub1tree.getroot()
        
        for sub1root in sub1tree.findall('name'):
            sheet.cell(sheetmaxrow3,15,sub1root.text)
        for sub1root in sub1tree.findall('type'):
            sheet.cell(sheetmaxrow3,16,sub1root.text)
        for sub1root in sub1tree.findall('handledBy'):
            sheet.cell(sheetmaxrow3,17,sub1root.text)
        for sub1root in sub1tree.findall('uri'):
            sheet.cell(sheetmaxrow3,18,sub1root.text)
        for params in sub1tree.iter('Param'):
            for pname in params.iter('name'):
                sheet.cell(sheetmaxrow3,19,pname.text)
            for pvalue in params.iter('value'):
                sheet.cell(sheetmaxrow3,20,pvalue.text)
            sheetmaxrow3=sheetmaxrow3+1      
    else:
        continue
lastrow=sheet.max_row

ws_to.cell(6,6).value=
    


wb.save('Finesse.xlsx')
