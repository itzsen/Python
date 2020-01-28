import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill,Border,Fill
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET

filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)


def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()


if os.path.isfile(filename):
    fileopen=open(filename)
    contents=fileopen.read()
    fileopen.close()
    ipRegex=re.compile(r'(ip:)(.*)')
    userRegex=re.compile(r'(username:)(.*)')
    passRegex=re.compile(r'(password:)(.*)')
    findip=ipRegex.search(contents)
    finduser=userRegex.search(contents)
    findpass=passRegex.search(contents)
    ip=findip.group(2)
    username=finduser.group(2)
    password=findpass.group(2)
else:
    fileopen=open(filename,'w')
    print('Config file does not exist. New config file created.')
    print('What is the Finesse FQDN address?')
    ip=input()
    newip='ip:'+ip+'\n'
    print('What is the Username?')
    username=input()
    newusername='username:'+username+'\n'
    print('What is the Password?')
    password=input()
    newpassword='password:'+password+'\n'
    fileopen.write(newip)
    fileopen.write(newusername)
    fileopen.write(newpassword)
    fileopen.close()
    
#URL
systeminfo='http://'+ip+'/finesse/api/SystemInfo'
systemconfig='http://'+ip+'/finesse/api/SystemConfig'
clusterconfig='http://'+ip+'/finesse/api/ClusterConfig'
enterpriseconfig='http://'+ip+'/finesse/api/EnterpriseDatabaseConfig'
layout='http://'+ip+'/finesse/api/LayoutConfig/default'
notready='http://'+ip+'/finesse/api/ReasonCodes?category=NOT_READY'
logout='http://'+ip+'/finesse/api/ReasonCodes?category=LOGOUT'
wrapup='http://'+ip+'/finesse/api/WrapUpReasons'
callvariable='http://'+ip+'/finesse/api/MediaPropertiesLayouts'
phonelayout='http://'+ip+'/finesse/api/PhoneBooks'
contactlayout=''
wflayout='http://'+ip+'/finesse/api/Workflows'
wfAlayout=''

# GET systeminfo
response=GET(systeminfo,username,password)
xmlcontent='systeminfo.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('systeminfo.xml')
root = tree.getroot()
for child in root:
    if child.tag=='currentTimestamp':
        currentTimestamp=child.text
    elif child.tag=='deploymentType':
        deploymentType=child.text
    elif child.tag=='lastCTIHeartbeatStatus':
        lastCTIHeartbeatStatus=child.text
    elif child.tag=='license':
        license1=child.text
    elif child.tag=='peripheralId':
        peripheralId=child.text
    elif child.tag=='status':
        status=child.text
    elif child.tag=='statusReason':
        statusReason=child.text
    elif child.tag=='systemAuthMode':
        systemAuthMode=child.text
    elif child.tag=='timezoneOffset':
        timezoneOffset=child.text
    elif child.tag=='uri':
        uri=child.text
    elif child.tag=='xmppDomain':
        xmppDomain=child.text
    elif child.tag=='xmppPubSubDomain':
        xmppPubSubDomain=child.text
    elif child.tag=='primaryNode':
        for element in child:
            primaryNode=element.text
    elif child.tag=='secondaryNode':
        for element in child:
            secondaryNode=element.text

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Cluster Setup')
sheet['B6']=currentTimestamp
sheet['B7']=deploymentType
sheet['B8']=lastCTIHeartbeatStatus
sheet['B9']=license1
sheet['B10']=peripheralId
sheet['B11']=primaryNode
sheet['B12']=secondaryNode
sheet['B13']=status
sheet['B14']=statusReason
sheet['B15']=systemAuthMode
sheet['B16']=timezoneOffset
sheet['B17']=xmppDomain
sheet['B18']=xmppPubSubDomain
sheet['B19']=uri
wb.save('Finesse.xlsx')

# GET systemconfig 
response=GET(systemconfig,username,password)
xmlcontent='systemconfig.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('systemconfig.xml')
root = tree.getroot()
for child in root:
    if child.tag=='uri':
        uri=child.text
    else:
        for element in child:
            if element.tag=='backupHost':
                backupHost=element.text
            elif element.tag=='backupPort':
                backupPort=element.text
            elif element.tag=='host':
                host=element.text
            elif element.tag=='peripheralId':
                peripheralId=element.text
            elif element.tag=='port':
                port=element.text

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Cluster Setup')
sheet['B22']=host
sheet['B23']=str(port)
sheet['B24']=backupHost
sheet['B25']=str(backupPort)
sheet['B26']=peripheralId
sheet['B27']=uri
wb.save('Finesse.xlsx')

# GET clusterconfig 
response=GET(clusterconfig,username,password)
xmlcontent='clusterconfig.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('clusterconfig.xml')
root = tree.getroot()
for child in root:
    if child.tag=='uri':
        uri=child.text
    else:
        for element in child:
            host=element.text

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Cluster Setup')
sheet['B30']=host
sheet['B31']=uri
wb.save('Finesse.xlsx')

# GET enterpriseconfig 
response=GET(enterpriseconfig,username,password)
xmlcontent='enterprisedatabaseconfig.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('enterprisedatabaseconfig.xml')
root = tree.getroot()
for child in root:
    if child.tag=='uri':
        uri=child.text
    elif child.tag=='backupHost':
        backupHost=child.text
    elif child.tag=='host':
        host=child.text
    elif child.tag=='databaseName':
        databaseName=child.text
    elif child.tag=='port':
        port=child.text
    elif child.tag=='password':
        ecpassword=child.text
    elif child.tag=='username':
        ecusername=child.text

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Cluster Setup')
sheet['B34']=host
sheet['B35']=backupHost
sheet['B36']=databaseName
sheet['B37']=port
sheet['B38']=ecusername
sheet['B39']=ecpassword
sheet['B40']=uri
ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')


# GET layout

response=GET(layout,username,password)
xmlcontent='layout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('layout.xml')
root = tree.getroot()
for child in root:
    if child.tag=='uri':
        uri=child.text
    elif child.tag=='layoutxml':
        layoutxml=child.text

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Layout')
sheet['B6']=layoutxml
sheet['B7']=uri

ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')

# GET notready
response=GET(notready,username,password)
xmlcontent='notready.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Not Ready')

i=7

tree= ET.parse('notready.xml')
root = tree.getroot()
for child in root.iter('ReasonCode'):
    for ruri in child.findall('uri'):
        sheet.cell(i,6,ruri.text)
    for rcategory in child.findall('category'):
        sheet.cell(i,1,rcategory.text)
    for rcode in child.findall('code'):
        sheet.cell(i,2,rcode.text)
    for rlabel in child.findall('label'):
        sheet.cell(i,3,rlabel.text)
    for rforall in child.findall('forAll'):
        sheet.cell(i,4,rforall.text)
    for rsystemcode in child.findall('systemCode'):
        sheet.cell(i,5,rsystemcode.text)
    i=i+1

ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')

# GET logout
response=GET(logout,username,password)
xmlcontent='logout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Logout')

i=7

tree= ET.parse('logout.xml')
root = tree.getroot()
for child in root.iter('ReasonCode'):
    for ruri in child.findall('uri'):
        sheet.cell(i,6,ruri.text)
    for rcategory in child.findall('category'):
        sheet.cell(i,1,rcategory.text)
    for rcode in child.findall('code'):
        sheet.cell(i,2,rcode.text)
    for rlabel in child.findall('label'):
        sheet.cell(i,3,rlabel.text)
    for rforall in child.findall('forAll'):
        sheet.cell(i,4,rforall.text)
    for rsystemcode in child.findall('systemCode'):
        sheet.cell(i,5,rsystemcode.text)
    i=i+1

ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')

# GET wrapup
response=GET(wrapup,username,password)
xmlcontent='wrapup.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Wrapup')

i=7

tree= ET.parse('wrapup.xml')
root = tree.getroot()
for child in root.iter('WrapUpReason'):
    for wlabel in child.findall('label'):
        sheet.cell(i,1,wlabel.text)
    for wforall in child.findall('forAll'):
        sheet.cell(i,2,wforall.text)
    for wuri in child.findall('uri'):
        sheet.cell(i,3,wuri.text)
    i=i+1

ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')

# GET callvariable

response=GET(callvariable,username,password)
xmlcontent='callvariable.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('callvariable.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Call Variable Layout')
sheetmaxrow=sheet.max_row+1
sheetmaxrow1=sheetmaxrow
sheetmaxrow2=sheetmaxrow

for child in root.findall('.//MediaPropertiesLayout//uri'):
    callvariableuri=child.text
    callvariableurivalue=callvariableuri.split('/')
    callvariableurl=('http://'+ip+'/finesse/api/MediaPropertiesLayout/'+callvariableurivalue[4])
    print ('######'+callvariableurl)

    response=GET(callvariableurl,username,password)
    xmlcontentfile=('callvariable'+callvariableurivalue[4])+'.xml'
    fileopen=open(xmlcontentfile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlcontentfile)
    subroot = subtree.getroot()
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,4,subchild.text)
    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,2,subchild.text)
    for subchild in subroot.findall('description'):
        sheet.cell(sheetmaxrow,3,subchild.text)
    for subchild in subroot.findall('type'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('./header/entry/displayName'):
        sheet.cell(sheetmaxrow,5,subchild.text)
    for subchild in subroot.findall('./header/entry/mediaProperty'):
        sheet.cell(sheetmaxrow,6,subchild.text)
    for subchild in subroot.findall('./column/entry/displayName'):
        sheet.cell(sheetmaxrow1,7,subchild.text)
        sheetmaxrow1=sheetmaxrow1+1
    for subchild in subroot.findall('./column/entry/mediaProperty'):
        sheet.cell(sheetmaxrow2,8,subchild.text)
        sheetmaxrow2=sheetmaxrow2+1
    sheetmaxrow=sheet.max_row+1
    sheetmaxrow1=sheetmaxrow
    sheetmaxrow2=sheetmaxrow

ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')

# GET phonelayout
response=GET(phonelayout,username,password)
xmlcontent='phonelayout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('phonelayout.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Phonebook')
sheetmaxrow=sheet.max_row+1

for child in root.findall('.//PhoneBook//uri'):
    phonebookuri=child.text
    phonebookurivalue=phonebookuri.split('/')
    phonebookurl=('http://'+ip+'/finesse/api/PhoneBook/'+phonebookurivalue[4])
    response=GET(phonebookurl,username,password)
    xmlcontentfile=('phonebook'+phonebookurivalue[4])+'.xml'
    fileopen=open(xmlcontentfile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlcontentfile)
    subroot = subtree.getroot()
    for subchild in subroot.findall('type'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,3,subchild.text)
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,4,subchild.text)
    for subchild in subroot.findall('contacts'):
        sheet.cell(sheetmaxrow,2,subchild.text)
        contactlayout='http://'+ip+subchild.text
        # GET contactlayout
        response=GET(contactlayout,username,password)
        cxmlcontent='contactlayout.xml'
        fileopen=open(cxmlcontent,'wb')
        fileopen.write(response.content)
        fileopen.close()

        ctree= ET.parse('contactlayout.xml')
        croot = ctree.getroot()
        sheetmaxrow1=sheetmaxrow
        sheetmaxrow2=sheetmaxrow
        sheetmaxrow3=sheetmaxrow
        sheetmaxrow4=sheetmaxrow
        sheetmaxrow5=sheetmaxrow
        for cchild in croot.findall('.//Contact//description'):
            sheet.cell(sheetmaxrow1,8,cchild.text)
            sheetmaxrow1=sheetmaxrow1+1
        for cchild in croot.findall('.//Contact//firstName'):
            sheet.cell(sheetmaxrow2,5,cchild.text)
            sheetmaxrow2=sheetmaxrow2+1
        for cchild in croot.findall('.//Contact//lastName'):
            sheet.cell(sheetmaxrow3,6,cchild.text)
            sheetmaxrow3=sheetmaxrow3+1
        for cchild in croot.findall('.//Contact//phoneNumber'):
            sheet.cell(sheetmaxrow4,7,cchild.text)
            sheetmaxrow4=sheetmaxrow4+1
        for cchild in croot.findall('.//Contact//uri'):
            sheet.cell(sheetmaxrow5,9,cchild.text)
            sheetmaxrow5=sheetmaxrow5+1
    sheetmaxrow=sheet.max_row+1

ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')

# GET workflowlayout
response=GET(wflayout,username,password)
xmlcontent='wflayout.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

tree= ET.parse('wflayout.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Workflow')
sheetmaxrow=sheet.max_row+1
sheetmaxrow1=sheetmaxrow
sheetmaxrow2=sheetmaxrow
sheetmaxrow3=sheetmaxrow


for child in root.findall('./Workflow/uri'):
    wfuri=child.text
    wfurivalue=wfuri.split('/')
    wfurl=('http://'+ip+'/finesse/api/Workflow/'+wfurivalue[4])
    response=GET(wfurl,username,password)
    xmlwffile=('workflow'+wfurivalue[4])+'.xml'
    fileopen=open(xmlwffile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlwffile)
    subroot = subtree.getroot()

    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('description'):
        sheet.cell(sheetmaxrow,2,subchild.text)
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,3,subchild.text)
    for subchild in subroot.findall('./TriggerSet/type'):
        sheet.cell(sheetmaxrow,4,subchild.text)
    for subchild in subroot.findall('./TriggerSet/name'):
        sheet.cell(sheetmaxrow,5,subchild.text)
    for subchild in subroot.findall('./ConditionSet/applyMethod'):
        sheet.cell(sheetmaxrow,6,subchild.text)
    for Condition in subroot.iter('Condition'):             
        for typec in Condition.iter('type'):
            sheet.cell(sheetmaxrow1,7,typec.text)
        for comparec in Condition.iter('comparator'):
            sheet.cell(sheetmaxrow1,9,comparec.text)
        for namec in Condition.iter('name'):
            sheet.cell(sheetmaxrow1,8,namec.text)
        for valuec in Condition.iter('value'):
            sheet.cell(sheetmaxrow1,10,valuec.text)      
        sheetmaxrow1=sheet.max_row+1

    for wfa in subroot.iter('WorkflowAction'):
        for wname in wfa.iter('name'):
            sheet.cell(sheetmaxrow2,11,wname.text)
        for wuri in wfa.iter('uri'):
            sheet.cell(sheetmaxrow2,12,wuri.text)         
        for wtype in wfa.iter('type'):
            sheet.cell(sheetmaxrow2,13,wtype.text) 
        sheetmaxrow2=sheetmaxrow2+1

    for subchild in subroot.findall('.//workflowActions//uri'):
        #GET workflow Action Parameters
        wfAuri='http://'+ip+subchild.text
        wfAurivalue=wfAuri.split('/')
        wfAurl=('http://'+ip+'/finesse/api/WorkflowAction/'+wfAurivalue[6])
        response=GET(wfAurl,username,password)
        wfAfile=('workflowaction'+wfAurivalue[6])+'.xml'
        fileopen=open(wfAfile,'wb')
        fileopen.write(response.content)
        fileopen.close()
   
    sheetmaxrow=sheet.max_row+2
    sheetmaxrow1=sheetmaxrow
    sheetmaxrow2=sheetmaxrow

for filenum in range (1,100):
    if os.path.isfile('workflowaction'+str(filenum)+'.xml'):
        wfactionfile='workflowaction'+str(filenum)+'.xml'
        sub1tree= ET.parse(wfactionfile)
        sub1root = sub1tree.getroot()
        
        for sub1root in sub1tree.findall('name'):
            sheet.cell(sheetmaxrow3,15,sub1root.text)
        for sub1root in sub1tree.findall('type'):
            sheet.cell(sheetmaxrow3,16,sub1root.text)
        for sub1root in sub1tree.findall('handledBy'):
            sheet.cell(sheetmaxrow3,17,sub1root.text)
        for sub1root in sub1tree.findall('uri'):
            sheet.cell(sheetmaxrow3,18,sub1root.text)
        for params in sub1tree.iter('Param'):
            for pname in params.iter('name'):
                sheet.cell(sheetmaxrow3,19,pname.text)
            for pvalue in params.iter('value'):
                sheet.cell(sheetmaxrow3,20,pvalue.text)
            sheetmaxrow3=sheetmaxrow3+1       
    else:
        continue

blackfill= PatternFill(start_color='FF000000',
                       end_color='FF000000', fill_type='solid')
lastrow=sheet.max_row

for cell in range(6,lastrow+1):
    sheet['N'+str(cell)].fill=blackfill

ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')


