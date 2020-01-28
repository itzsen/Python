import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill,Border,Fill
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET

filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

ip='10.2.2.35'
username='ucadmin'
password='Syr!nx2112'

def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()

notready='http://'+ip+'/finesse/api/ReasonCodes?category=NOT_READY'

wrapup='http://'+ip+'/finesse/api/WrapUpReasons'

# GET notready

response=GET(wrapup,username,password)
xmlcontent='wrapup.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Wrapup')


i=7

tree= ET.parse('wrapup.xml')
root = tree.getroot()
for child in root.iter('WrapUpReason'):
    for wlabel in child.findall('label'):
        sheet.cell(i,1,wlabel.text)
    for wforall in child.findall('forAll'):
        sheet.cell(i,2,wforall.text)
    for wuri in child.findall('uri'):
        sheet.cell(i,3,wuri.text)
    i=i+1


ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')


sheet=wb.get_sheet_by_name('Title')
ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')
