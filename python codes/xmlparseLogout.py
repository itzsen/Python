import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill,Border,Fill
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET

filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

ip='10.2.2.35'
username='ucadmin'
password='Syr!nx2112'

def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()

notready='http://'+ip+'/finesse/api/ReasonCodes?category=NOT_READY'



# GET notready

response=GET(notready,username,password)
xmlcontent='notready.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()

wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Logout')


i=7

tree= ET.parse('logout.xml')
root = tree.getroot()
for child in root.iter('ReasonCode'):
    for ruri in child.findall('uri'):
        sheet.cell(i,6,ruri.text)
    for rcategory in child.findall('category'):
        sheet.cell(i,1,rcategory.text)
    for rcode in child.findall('code'):
        sheet.cell(i,2,rcode.text)
    for rlabel in child.findall('label'):
        sheet.cell(i,3,rlabel.text)
    for rforall in child.findall('forAll'):
        sheet.cell(i,4,rforall.text)
    for rsystemcode in child.findall('systemCode'):
        sheet.cell(i,5,rsystemcode.text)
    i=i+1


ws=wb.active
img=Image('DDlogo.png')
ws.add_image (img, 'A1')
wb.save('Finesse.xlsx')



