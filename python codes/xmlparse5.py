import os,re,sys,shutil

# Using Python external "requests" module to call REST APIs
import requests
import openpyxl
from openpyxl.drawing.image import Image
from requests.auth import HTTPBasicAuth
from xml.etree import ElementTree as ET


filename='config.txt'
templateExcel='template.xlsx'
finalExcel='Finesse.xlsx'
shutil.copy(templateExcel,finalExcel)

username='ucadmin'
password='Syr!nx2112'
ip='10.2.2.35'


def printRESTAPIResponse(prefix='printRESTAPIResponse()', response=''):
    """
        Print the REST API response with the HTTP Request status code
        and the Response body text
    """
    print (prefix + " - Status: ", response.status_code) # This is the http request status
    print (prefix + " - Text: ", response.text)
    print ("\n")

def GET(url, username, password, params=''):
    """
        Call the GET HTTP Request using HTTP Basic Auth authentication
        
        Parameters:
            url (str): The URL to make the REST request
            username (str): The username of the user making the HTTP request
            password (str): The password of the user making the HTTP request
            params(dictionary, optional): Dictionary or bytes to be sent in the query string for the Request.
                                          (e.g. {"category" : "NOT_READY"})

        Returns: Response object (http://docs.python-requests.org/en/master/api/#requests.Response) - The HTTP Response as a result of the HTTP Request
    """
    print ("Executing GET '%s'" % url)
    try:
        response = requests.get(url=url, auth=HTTPBasicAuth(username, password), params=params)
        printRESTAPIResponse("GET()", response)
        return(response)
    except:
       print ("An error occured in the GET request to %s" % url)
       print (sys.exc_info());
       sys.exit()



callvariable='http://10.2.2.35/finesse/api/MediaPropertiesLayouts'

# GET callvariable

response=GET(callvariable,username,password)
xmlcontent='callvariable.xml'
fileopen=open(xmlcontent,'wb')
fileopen.write(response.content)
fileopen.close()


tree= ET.parse('callvariable.xml')
root = tree.getroot()
wb=openpyxl.load_workbook('Finesse.xlsx')
sheet=wb.get_sheet_by_name('Call Variable Layout')
sheetmaxrow=sheet.max_row+1
sheetmaxrow1=sheetmaxrow
sheetmaxrow2=sheetmaxrow

for child in root.findall('.//MediaPropertiesLayout//uri'):
    callvariableuri=child.text
    callvariableurivalue=callvariableuri.split('/')
    callvariableurl=('http://'+ip+'/finesse/api/MediaPropertiesLayout/'+callvariableurivalue[4])
    print ('######'+callvariableurl)

    response=GET(callvariableurl,username,password)
    xmlcontentfile=('callvariable'+callvariableurivalue[4])+'.xml'
    fileopen=open(xmlcontentfile,'wb')
    fileopen.write(response.content)
    fileopen.close()

    subtree= ET.parse(xmlcontentfile)
    subroot = subtree.getroot()
    for subchild in subroot.findall('uri'):
        sheet.cell(sheetmaxrow,4,subchild.text)
    for subchild in subroot.findall('name'):
        sheet.cell(sheetmaxrow,2,subchild.text)
    for subchild in subroot.findall('description'):
        sheet.cell(sheetmaxrow,3,subchild.text)
    for subchild in subroot.findall('type'):
        sheet.cell(sheetmaxrow,1,subchild.text)
    for subchild in subroot.findall('./header/entry/displayName'):
        sheet.cell(sheetmaxrow,5,subchild.text)
    for subchild in subroot.findall('./header/entry/mediaProperty'):
        sheet.cell(sheetmaxrow,6,subchild.text)
    for subchild in subroot.findall('./column/entry/displayName'):
        sheet.cell(sheetmaxrow1,7,subchild.text)
        sheetmaxrow1=sheetmaxrow1+1
    for subchild in subroot.findall('./column/entry/mediaProperty'):
        sheet.cell(sheetmaxrow2,8,subchild.text)
        sheetmaxrow2=sheetmaxrow2+1
    sheetmaxrow=sheet.max_row+1
    sheetmaxrow1=sheetmaxrow
    sheetmaxrow2=sheetmaxrow
wb.save('Finesse.xlsx')


